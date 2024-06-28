
from openpyxl import load_workbook
import re
from openpyxl import Workbook

def parse_header(header_file):
    with open(header_file, 'r') as f:
        content = f.read()

   
    pattern = r'\b([\w*]+)\s+(\w+)\s*\((.*?)\);'

    
    prototypes = re.findall(pattern, content)

    return prototypes

def print_excel_contents(excel_file):
    wb = load_workbook(excel_file)
    sheet = wb.active

    print(f"Contents of '{excel_file}':")
    for row in sheet.iter_rows(values_only=True):
        print(row)

    wb.close()

def main():
    header_file = 'headers.h' 
    prototypes = parse_header(header_file)

    # Create Excel workbook and sheet
    wb = Workbook()
    sheet = wb.active
    sheet.title = 'Function Prototypes'

    # Write headers
    sheet['A1'] = 'ID'
    sheet['B1'] = 'Return Type'
    sheet['C1'] = 'Function Name'
    sheet['D1'] = 'Arguments'

    
    for idx, (return_type, function_name, arguments) in enumerate(prototypes, start=1):
        unique_id = f"IDX{idx}"
        sheet.append([unique_id, return_type, function_name, arguments])

   
    excel_file = 'function_prototypes.xlsx'  
    wb.save(excel_file)
    print(f"Function prototypes written to '{excel_file}'.")

   
    print_excel_contents(excel_file)


main()
